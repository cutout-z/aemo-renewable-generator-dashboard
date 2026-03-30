"""Excel output generation with per-state workbooks."""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from . import config

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_all_workbooks(summary: pd.DataFrame, output_dir: str):
    """Generate one .xlsx workbook per state from the summary DataFrame."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    for region in config.REGIONS:
        friendly = config.REGION_NAMES[region]
        region_data = summary[summary["REGIONID"] == region].copy()
        if region_data.empty:
            logger.warning(f"No data for {friendly}, skipping workbook")
            continue

        filepath = output_path / f"{friendly}_curtailment.xlsx"
        _write_region_workbook(region_data, friendly, summary.columns.tolist(), filepath)
        logger.info(f"Written {filepath}")


def _write_region_workbook(data: pd.DataFrame, region_name: str,
                           all_columns: list[str], filepath: Path):
    """Write a workbook for a single region."""
    wb = Workbook()

    _write_summary_table(wb, data, region_name, all_columns)
    _write_heatmap(wb, data, region_name)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)


def _write_summary_table(wb: Workbook, data: pd.DataFrame, region_name: str,
                          all_columns: list[str]):
    """Sheet 1: Full summary table."""
    ws = wb.create_sheet(title="Summary")

    # Define column order and labels
    col_spec = _get_column_spec(data)

    headers = [label for _, label in col_spec]
    col_keys = [key for key, _ in col_spec]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Sort by fuel type then project name
    sort_cols = [c for c in ["FUEL_TYPE", "PROJECT_NAME"] if c in data.columns]
    if sort_cols:
        data = data.sort_values(sort_cols)

    # Write data
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        for col_idx, key in enumerate(col_keys, 1):
            val = row.get(key)
            if pd.isna(val):
                cell_val = ""
            elif isinstance(val, float):
                cell_val = val
            else:
                cell_val = val

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

            # Number formatting
            if key.startswith(("MLF_", "ELI_", "CURTAILMENT_ACTUAL_")):
                cell.number_format = "0.0000"
            elif key.startswith(("ISP_CURTAILMENT_", "ISP_OFFLOADING_")):
                cell.number_format = "0.00"
            elif key == "NAMEPLATE_MW":
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right")

            # Left-align text columns
            if key in ("PROJECT_NAME", "LOCATION", "REZ_NAME"):
                cell.alignment = Alignment(horizontal="left")

    # Column widths
    widths = {
        "DUID": 14, "PROJECT_NAME": 30, "FUEL_TYPE": 8, "LOCATION": 16,
        "REZ": 5, "REZ_NAME": 22, "STATE": 6, "NAMEPLATE_MW": 12,
        "VOLTAGE_KV": 10,
    }
    for col_idx, key in enumerate(col_keys, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(key, 12)

    ws.freeze_panes = "C2"


def _write_heatmap(wb: Workbook, data: pd.DataFrame, region_name: str):
    """Sheet 2: Curtailment + MLF values with conditional colour formatting."""
    ws = wb.create_sheet(title="Heatmap")

    # Get numeric columns for heatmap
    value_cols = [c for c in data.columns if c.startswith(("MLF_", "ELI_", "CURTAILMENT_ACTUAL_", "ISP_"))]
    if not value_cols:
        ws.cell(row=1, column=1, value="No numeric data available")
        return

    headers = ["DUID", "Fuel"] + value_cols
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    num_rows = len(data)
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row.get("DUID", "")).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=row.get("FUEL_TYPE", "")).border = THIN_BORDER
        for col_offset, col in enumerate(value_cols):
            val = row.get(col)
            cell = ws.cell(row=row_idx, column=3 + col_offset,
                           value=val if pd.notna(val) else "")
            cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Apply colour scales
    if num_rows > 0:
        for col_idx, col in enumerate(value_cols, 3):
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{num_rows + 1}"

            if col.startswith("MLF_"):
                # MLF: red (low) → yellow → green (high) — same as MLF tracker
                ws.conditional_formatting.add(
                    cell_range,
                    ColorScaleRule(
                        start_type="min", start_color="F8696B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="63BE7B",
                    ),
                )
            else:
                # Curtailment/offloading: green (low = good) → yellow → red (high = bad)
                ws.conditional_formatting.add(
                    cell_range,
                    ColorScaleRule(
                        start_type="min", start_color="63BE7B",
                        mid_type="percentile", mid_value=50, mid_color="FFEB84",
                        end_type="max", end_color="F8696B",
                    ),
                )

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    for i in range(3, 3 + len(value_cols)):
        ws.column_dimensions[get_column_letter(i)].width = 14

    ws.freeze_panes = "C2"


def _get_column_spec(data: pd.DataFrame) -> list[tuple[str, str]]:
    """Build ordered column specification (key, label) based on available data."""
    spec = [
        ("DUID", "DUID"),
        ("PROJECT_NAME", "Project Name"),
        ("FUEL_TYPE", "Fuel"),
        ("LOCATION", "Location"),
        ("REZ", "REZ"),
        ("REZ_NAME", "REZ Name"),
        ("STATE", "State"),
        ("NAMEPLATE_MW", "MW"),
        ("VOLTAGE_KV", "kV"),
    ]

    # Add actual curtailment columns
    for col in sorted(data.columns):
        if col.startswith("CURTAILMENT_ACTUAL_"):
            label = col.replace("CURTAILMENT_ACTUAL_", "Actual ")
            spec.append((col, label))

    # Add ELI projected curtailment
    if "ELI_CURTAILMENT_NEAR" in data.columns:
        spec.append(("ELI_CURTAILMENT_NEAR", "ELI Near Term"))
    if "ELI_CURTAILMENT_MED" in data.columns:
        spec.append(("ELI_CURTAILMENT_MED", "ELI Med Term"))

    # Add MLF columns
    for col in sorted(data.columns):
        if col.startswith("MLF_"):
            label = col.replace("MLF_", "")
            spec.append((col, label))

    # Add ISP forecast columns
    for col in sorted(data.columns):
        if col.startswith("ISP_CURTAILMENT_FY") and not col.endswith("_LABEL"):
            label = col.replace("ISP_CURTAILMENT_", "Curt ")
            spec.append((col, label))
    if "ISP_CURTAILMENT_AVG" in data.columns:
        spec.append(("ISP_CURTAILMENT_AVG", "Curt Avg"))

    for col in sorted(data.columns):
        if col.startswith("ISP_OFFLOADING_FY") and not col.endswith("_LABEL"):
            label = col.replace("ISP_OFFLOADING_", "Offl ")
            spec.append((col, label))
    if "ISP_OFFLOADING_AVG" in data.columns:
        spec.append(("ISP_OFFLOADING_AVG", "Offl Avg"))

    # Filter to columns that actually exist in data
    return [(key, label) for key, label in spec if key in data.columns]
